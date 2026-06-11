from __future__ import annotations

from datetime import date
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app import db
from app.main import app


def make_workbook(rows: list[list[object]], meta: dict[str, object] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quote"
    for row in rows:
        sheet.append(row)
    if meta:
        meta_sheet = workbook.create_sheet("Meta", 0)
        for key, value in meta.items():
            meta_sheet.append([key, value])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_import_quotation_and_catalog_history(tmp_path, monkeypatch):
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "quotations.db")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 'quotations.db'}")
    monkeypatch.setenv("STORAGE_BACKEND", "local")
    monkeypatch.setenv("LOCAL_STORAGE_DIR", str(tmp_path / "uploads"))
    db.init_db()
    client = TestClient(app)

    headers = ["catalogNo", "Product Name", "Unit Price", "Qty", "Unit", "Notes"]
    first = make_workbook(
        [
            headers,
            ["ACS6001-10", "Example Product", "120.50", 10, "g", "first quote"],
            [None, "Missing catalog", 8, 1, "g", "skipped"],
        ],
        {"Customer": "Acme Pharma", "Quote Date": "2025-03-01", "Currency": "USD"},
    )
    second = make_workbook([headers, ["ACS6001-10", "Example Product", "115.00", 10, "g", "discounted"]])

    response = client.post("/api/quotations/import", files={"file": ("q1.xlsx", first)}, data={"currency": "USD"})
    assert response.status_code == 200
    assert response.json()["imported_rows"] == 1
    assert response.json()["skipped_rows"] == 1
    assert response.json()["stored_file"]["sha256"]
    first_id = response.json()["quotation_id"]

    response = client.post(
        "/api/quotations/import",
        files={"file": ("q2.xlsx", second)},
        data={"customer": "Acme Pharma", "quoted_on": "2025-06-01", "currency": "USD"},
    )
    assert response.status_code == 200

    history = client.get("/api/history/ACS6001-10")
    assert history.status_code == 200
    points = history.json()["history"]
    assert [point["unit_price"] for point in points] == [120.5, 115.0]
    assert [point["quoted_on"] for point in points] == ["2025-03-01", "2025-06-01"]

    detail = client.get(f"/api/quotations/{first_id}")
    assert detail.status_code == 200
    assert detail.json()["quotation"]["line_item_count"] == 1
    assert detail.json()["quotation"]["line_items"][0]["catalog_no"] == "ACS6001-10"

    listing = client.get("/api/quotations")
    assert listing.status_code == 200
    assert len(listing.json()["quotations"]) == 2


def test_import_quotation_rejects_missing_catalog_column(tmp_path, monkeypatch):
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "quotations.db")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 'quotations.db'}")
    db.init_db()
    client = TestClient(app)
    workbook = make_workbook([["Product Name", "Unit Price"], ["Example", 10]])

    response = client.post("/api/quotations/import", files={"file": ("bad.xlsx", workbook)})

    assert response.status_code == 400
    assert "catalog number" in response.json()["detail"].lower()


def test_default_users_login_and_archive_quotation(tmp_path, monkeypatch):
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "quotations.db")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 'quotations.db'}")
    monkeypatch.setenv("STORAGE_BACKEND", "local")
    monkeypatch.setenv("LOCAL_STORAGE_DIR", str(tmp_path / "uploads"))
    db.init_db()
    client = TestClient(app)

    login = client.post("/api/auth/login", json={"email": "admin@example.com", "password": "Admin@123456"})
    assert login.status_code == 200
    assert login.json()["user"]["role"] == "admin"
    token = login.json()["token"]

    user_login = client.post("/api/auth/login", json={"email": "user@example.com", "password": "User@123456"})
    assert user_login.status_code == 200
    assert user_login.json()["user"]["role"] == "user"

    workbook = make_workbook([["catalogNo", "Product Name", "Unit Price"], ["ACS6001-10", "Example", 12]])
    imported = client.post("/api/quotations/import", files={"file": ("q.xlsx", workbook)})
    quotation_id = imported.json()["quotation_id"]

    archived = client.delete(f"/api/quotations/{quotation_id}", headers={"Authorization": f"Bearer {token}"})
    assert archived.status_code == 200

    listing = client.get("/api/quotations")
    assert listing.json()["quotations"] == []


def test_manual_record_requires_login_and_uses_current_user_and_date(tmp_path, monkeypatch):
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "records.db")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 'records.db'}")
    db.init_db()
    client = TestClient(app)
    payload = {
        "material_name": "ethanol",
        "price": "12.345",
        "quantity": "100g",
        "record_type": "inquiry",
        "record_date": "1999-01-01",
        "requester": "spoofed",
    }

    denied = client.post("/api/records", json=payload)
    assert denied.status_code == 401

    login = client.post("/api/auth/login", json={"email": "user@example.com", "password": "User@123456"})
    token = login.json()["token"]
    created = client.post("/api/records", json=payload, headers={"Authorization": f"Bearer {token}"})
    assert created.status_code == 200

    with db.connect() as conn:
        row = db.row_to_dict(conn.execute("select * from procurement_records where id = ?", (created.json()["id"],)).fetchone())
    assert row["requester"] == "User"
    assert row["record_date"] == date.today().isoformat()
    assert row["price"] == 12.35 or row["price"] == "12.35"
    assert row["unit_price"] == 0.12 or row["unit_price"] == "0.12"


def test_admin_import_confirm_records_current_user_and_batch_quote(tmp_path, monkeypatch):
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "import.db")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 'import.db'}")
    monkeypatch.setenv("STORAGE_BACKEND", "local")
    monkeypatch.setenv("LOCAL_STORAGE_DIR", str(tmp_path / "uploads"))
    db.init_db()
    client = TestClient(app)
    login = client.post("/api/auth/login", json={"email": "admin@example.com", "password": "Admin@123456"})
    token = login.json()["token"]
    headers = {"Authorization": f"Bearer {token}"}
    content = "化合物名称,价格,采购数量,供应商\n乙醇,24.789,100g,Acme\n".encode()

    denied = client.post("/api/import/preview", files={"file": ("records.csv", content, "text/csv")})
    assert denied.status_code == 403

    preview = client.post("/api/import/preview", files={"file": ("records.csv", content, "text/csv")}, headers=headers)
    assert preview.status_code == 200
    assert preview.json()["preview_rows"][0]["normalized"]["price"] == "24.79"

    confirm = client.post(
        "/api/import/confirm",
        json={
            "filename": preview.json()["filename"],
            "import_type": "purchase",
            "rows": preview.json()["preview_rows"],
            "mapping": preview.json()["mapping"],
            "uploaded_by": "spoofed",
            "stored_file_id": preview.json()["stored_file_id"],
        },
        headers=headers,
    )
    assert confirm.status_code == 200

    with db.connect() as conn:
        import_file = db.row_to_dict(conn.execute("select * from import_files order by id desc limit 1").fetchone())
        record = db.row_to_dict(conn.execute("select * from procurement_records order by id desc limit 1").fetchone())
    assert import_file["uploaded_by"] == "admin@example.com"
    assert record["created_by"] == "admin@example.com"
    assert record["price"] == 24.79 or record["price"] == "24.79"

    quote = client.post(
        "/api/quotations/calculate-batch",
        json={
            "items": [
                {"material_query": "ethanol", "required_quantity": "50g", "required_unit": "g"},
                {"material_query": "missing-material", "required_quantity": "1g", "required_unit": "g"},
            ]
        },
    )
    assert quote.status_code == 200
    items = quote.json()["items"]
    assert items[0]["result"]["estimated_cost"] == "12.50"
    assert items[1]["result"]["warning"] == "未找到匹配物料。"

    dashboard = client.get("/api/analytics/dashboard")
    assert dashboard.status_code == 200
    data = dashboard.json()
    assert data["record_count"] == 1
    assert data["most_recorded_materials"][0]["standard_name"] == "Ethanol"
    assert data["supplier_summary"][0]["supplier"] == "Acme"


def test_register_structure_and_link_to_material(tmp_path, monkeypatch):
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "structures.db")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 'structures.db'}")
    db.init_db()
    client = TestClient(app)
    login = client.post("/api/auth/login", json={"email": "user@example.com", "password": "User@123456"})
    token = login.json()["token"]
    headers = {"Authorization": f"Bearer {token}"}

    denied = client.post("/api/structures", json={"name": "Ethanol", "smiles": "CCO"})
    assert denied.status_code == 401

    created = client.post(
        "/api/structures",
        json={"name": "Ethanol", "cas_number": "64-17-5", "smiles": "CCO", "notes": "registered from editor"},
        headers=headers,
    )
    assert created.status_code == 200
    structure = created.json()["structure"]
    assert structure["name"] == "Ethanol"
    assert structure["canonical_smiles"]
    assert structure["structure_svg"]

    duplicate = client.post("/api/structures", json={"name": "Ethanol duplicate", "smiles": "CCO"}, headers=headers)
    assert duplicate.status_code == 409

    listing = client.get("/api/structures?q=ethanol")
    assert listing.status_code == 200
    assert listing.json()["structures"][0]["name"] == "Ethanol"

    detail = client.get(f"/api/materials/{structure['material_id']}/records")
    assert detail.status_code == 200
    assert detail.json()["structures"][0]["id"] == structure["id"]


def test_isislike_molecule_compatibility_endpoints(tmp_path, monkeypatch):
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "molecules.db")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 'molecules.db'}")
    db.init_db()
    client = TestClient(app)

    saved = client.post("/api/molecules/save", json={"smiles": "CCO", "molfile": None})
    assert saved.status_code == 200
    molecule = saved.json()
    assert molecule["id"]
    assert molecule["canonical_smiles"]

    duplicate = client.post("/api/molecules/save", json={"smiles": "CCO", "molfile": None})
    assert duplicate.status_code == 200
    assert duplicate.json()["id"] == molecule["id"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name", "smiles", "notes"])
    sheet.append(["Acetone", "CC(=O)C", "from isislike import"])
    buffer = BytesIO()
    workbook.save(buffer)
    imported = client.post(
        "/api/molecules/import",
        files={"file": ("molecules.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 200
    assert imported.json()["success_count"] == 1
    assert imported.json()["failed_count"] == 0

    listing = client.get("/api/molecules")
    assert listing.status_code == 200
    assert len(listing.json()) == 2

    exact = client.post("/api/molecules/search/exact", json={"smiles": "CCO"})
    assert exact.status_code == 200
    assert exact.json()["id"] == molecule["id"]

    svg = client.get(f"/api/molecules/{molecule['id']}/structure.svg")
    assert svg.status_code == 200
    assert "svg" in svg.text
    acetone = client.post("/api/molecules/search/exact", json={"smiles": "CC(=O)C"}).json()
    acetone_svg = client.get(f"/api/molecules/{acetone['id']}/structure.svg")
    assert acetone_svg.status_code == 200
    assert acetone_svg.text != svg.text

    detail = client.get(f"/api/molecules/{molecule['id']}")
    assert detail.status_code == 200
    assert detail.json()["has_structure_svg"] is True

    updated = client.patch(
        f"/api/molecules/{molecule['id']}",
        json={"name": "Ethanol updated", "notes": "lab batch"},
    )
    assert updated.status_code == 200
    assert updated.json()["name"] == "Ethanol updated"
    assert updated.json()["notes"] == "lab batch"

    deleted = client.delete(f"/api/molecules/{molecule['id']}")
    assert deleted.status_code == 204
    assert client.get(f"/api/molecules/{molecule['id']}").status_code == 404

    config = client.get("/api/export/config")
    assert config.status_code == 200
    assert config.json()["enabled"] is False
