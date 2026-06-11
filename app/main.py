from __future__ import annotations

import csv
import io
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, Form, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from app import db
from app import auth
from app.config import get_settings
from app.services.chemical_resolver import resolve_chemical
from app.services.importer import build_preview
from app.services.pdf_parser import extract_text_pdf
from app.services.quotation_excel import parse_quotation_workbook
from app.services.quotation_repository import (
    archive_quotation,
    compare_quotations,
    create_quotation,
    get_catalog_history,
    get_quotation_detail,
    get_quotation_summary,
    list_quotations,
    quotation_dashboard,
)
from app.services.quote_engine import choose_price_source, estimate_cost
from app.services.structure_registry import (
    StructureError,
    archive_structure,
    derive_structure_properties,
    get_structure,
    list_structures,
    register_structure,
    render_smiles_svg,
    update_structure,
)
from app.storage import get_storage
from app.services.units import calculate_unit_price, parse_decimal, parse_quantity, quantize_money

settings = get_settings()
app = FastAPI(title=settings.app_name)
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

STATIC_DIR = Path(__file__).resolve().parent.parent / "static"
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
ISISLIKE_DIST_DIR = STATIC_DIR / "isislike"
if ISISLIKE_DIST_DIR.exists():
    app.mount("/isislike", StaticFiles(directory=ISISLIKE_DIST_DIR, html=True), name="isislike")


class ManualRecord(BaseModel):
    material_name: str = Field(min_length=1)
    price: Decimal
    currency: str = "CNY"
    quantity: str | Decimal | None = None
    unit: str | None = None
    record_type: str = Field(pattern="^(inquiry|purchase)$")
    record_date: str | None = None
    requester: str | None = None
    supplier: str | None = None
    remark: str | None = None
    cas_number: str | None = None


class ConfirmImport(BaseModel):
    filename: str
    import_type: str = Field(pattern="^(inquiry|purchase)$")
    rows: list[dict[str, Any]]
    mapping: dict[str, str] = {}
    uploaded_by: str | None = None
    stored_file_id: int | None = None


class QuoteRequest(BaseModel):
    material_query: str
    required_quantity: str | Decimal
    required_unit: str | None = None


class QuoteBatchRequest(BaseModel):
    items: list[QuoteRequest] = Field(min_length=1)


class RegisterRequest(BaseModel):
    email: str
    password: str = Field(min_length=8)
    name: str | None = None


class LoginRequest(BaseModel):
    email: str
    password: str


class FxRatePayload(BaseModel):
    base_currency: str
    quote_currency: str
    rate: Decimal
    effective_on: str


class StructureRegisterRequest(BaseModel):
    name: str = Field(min_length=1)
    smiles: str | None = None
    molfile: str | None = None
    cas_number: str | None = None
    notes: str | None = None
    material_id: int | None = None


class MoleculeSaveRequest(BaseModel):
    smiles: str
    molfile: str | None = None
    name: str | None = None
    notes: str | None = None


class MoleculeUpdateRequest(BaseModel):
    name: str | None = None
    notes: str | None = None


class SmilesInput(BaseModel):
    smiles: str


class SmartsInput(BaseModel):
    smarts: str


class SimilaritySearchRequest(BaseModel):
    smiles: str
    match_threshold: float = 0.7
    match_count: int = 50


def _clean_text(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def get_request_user(authorization: str | None, x_user_role: str | None) -> dict | None:
    return auth.current_user(authorization=authorization, x_user_role=x_user_role)


def require_admin(authorization: str | None, role: str | None) -> dict | None:
    user = get_request_user(authorization, role)
    auth.require_role(user, "admin")
    return user


def require_login(authorization: str | None, role: str | None) -> dict:
    user = get_request_user(authorization, role)
    auth.require_authenticated(user)
    return user or {}


def decimal_to_str(value: Decimal | None) -> str | None:
    return str(value) if value is not None else None


def remember_stored_file(filename: str, content: bytes, content_type: str | None, created_by: int | None = None) -> dict:
    stored = get_storage().put(filename, content, content_type)
    with db.connect() as conn:
        duplicate_row = conn.execute("select id from stored_files where sha256 = ? limit 1", (stored.sha256,)).fetchone()
        cur = conn.execute(
            """
            insert into stored_files (
                original_filename, storage_backend, bucket, object_key, sha256, mime_type, size_bytes, created_by
            )
            values (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (filename, stored.storage_backend, stored.bucket, stored.key, stored.sha256, content_type, stored.size, created_by),
        )
    return {
        "id": int(cur.lastrowid),
        "original_filename": filename,
        "storage_backend": stored.storage_backend,
        "bucket": stored.bucket,
        "object_key": stored.key,
        "sha256": stored.sha256,
        "size_bytes": stored.size,
        "is_duplicate": duplicate_row is not None,
    }


@app.on_event("startup")
def startup() -> None:
    db.init_db()


@app.get("/", response_class=HTMLResponse)
def index() -> str:
    return (STATIC_DIR / "index.html").read_text(encoding="utf-8")


@app.get("/api/health")
def health() -> dict:
    return {"status": "ok", "environment": settings.environment, "storage_backend": settings.storage_backend}


@app.get("/health")
def root_health() -> dict:
    return health()


@app.post("/api/quotations/import")
async def import_quotation(
    file: UploadFile = File(...),
    customer: str | None = Form(default=None),
    quoted_on: str | None = Form(default=None),
    currency: str | None = Form(default=None),
) -> dict:
    filename = file.filename or "quotation.xlsx"
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx and .xlsm quotation files are supported.")

    content = await file.read()
    if len(content) > settings.max_upload_mb * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Uploaded file is too large.")
    try:
        parsed = parse_quotation_workbook(content)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Unable to parse quotation workbook.") from exc

    if not parsed.line_items:
        raise HTTPException(status_code=400, detail="No importable quotation line items found.")

    stored = remember_stored_file(filename, content, file.content_type)
    quotation_id = create_quotation(filename, parsed, customer, quoted_on, currency, stored["id"], stored["sha256"])
    quotation = get_quotation_summary(quotation_id)
    return {
        "quotation_id": quotation_id,
        "imported_rows": len(parsed.line_items),
        "skipped_rows": parsed.skipped_rows,
        "stored_file": stored,
        "quotation": quotation,
    }


@app.get("/api/quotations")
def quotation_list() -> dict:
    return {"quotations": list_quotations()}


@app.get("/api/quotations/{quotation_id}")
def quotation_detail(quotation_id: int) -> dict:
    quotation = get_quotation_detail(quotation_id)
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found.")
    return {"quotation": quotation}


@app.get("/api/history/{catalog_no}")
def catalog_history(catalog_no: str) -> dict:
    history = get_catalog_history(catalog_no)
    if not history:
        raise HTTPException(status_code=404, detail="No quotation history found for catalog number.")
    return {"catalog_no": catalog_no, "history": history}


@app.post("/api/auth/register")
def register(payload: RegisterRequest, authorization: str | None = Header(default=None), x_user_role: str | None = Header(default=None)) -> dict:
    db.init_db()
    with db.connect() as conn:
        has_users = conn.execute("select id from users limit 1").fetchone() is not None
    role = "admin" if not has_users else "user"
    if has_users:
        require_admin(authorization, x_user_role)
    user = auth.create_user(payload.email, payload.password, payload.name, role)
    token = auth.issue_token(int(user["id"]))
    return {"user": user, "token": token}


@app.post("/api/auth/login")
def login(payload: LoginRequest) -> dict:
    user = auth.authenticate(payload.email, payload.password)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password.")
    return {"user": user, "token": auth.issue_token(int(user["id"]))}


@app.get("/api/auth/me")
def me(authorization: str | None = Header(default=None), x_user_role: str | None = Header(default=None)) -> dict:
    user = get_request_user(authorization, x_user_role)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated.")
    return {"user": user}


@app.post("/api/import/preview")
async def import_preview(file: UploadFile = File(...), authorization: str | None = Header(default=None), x_user_role: str | None = Header(default=None)) -> dict:
    user = require_admin(authorization, x_user_role)
    content = await file.read()
    if len(content) > settings.max_upload_mb * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Uploaded file is too large.")
    stored = remember_stored_file(file.filename or "upload.xlsx", content, file.content_type, user.get("id") if user else None)
    try:
        preview = await build_preview(file.filename or "upload.xlsx", content)
        preview["stored_file"] = stored
        preview["stored_file_id"] = stored["id"]
        preview["is_duplicate"] = stored["is_duplicate"]
        return preview
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/import/confirm")
async def import_confirm(payload: ConfirmImport, authorization: str | None = Header(default=None), x_user_role: str | None = Header(default=None)) -> dict:
    user = require_admin(authorization, x_user_role)
    uploaded_by = str(user.get("email") or user.get("name") or user.get("id"))
    db.init_db()
    success = 0
    failed = 0
    with db.connect() as conn:
        stored = None
        if payload.stored_file_id:
            stored = db.row_to_dict(conn.execute("select * from stored_files where id = ?", (payload.stored_file_id,)).fetchone())
        duplicate = 0
        if stored:
            existing = conn.execute("select id from import_files where file_sha256 = ? and archived_at is null limit 1", (stored["sha256"],)).fetchone()
            duplicate = 1 if existing else 0
        cur = conn.execute(
            """
            insert into import_files (filename, uploaded_by, import_type, status, total_rows, stored_file_id, file_sha256, is_duplicate)
            values (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (payload.filename, uploaded_by, payload.import_type, "confirmed", len(payload.rows), payload.stored_file_id, stored["sha256"] if stored else None, duplicate),
        )
        file_id = int(cur.lastrowid)

        for header, field in payload.mapping.items():
            conn.execute(
                """
                insert into field_mappings (excel_header, system_field, confirmed_by, updated_at)
                values (?, ?, ?, current_timestamp)
                on conflict(excel_header) do update set system_field = excluded.system_field, updated_at = current_timestamp
                """,
                (header, field, uploaded_by),
            )

    for item in payload.rows:
        normalized = item.get("normalized", item)
        try:
            await create_record_from_normalized(normalized, payload.import_type, uploaded_by, file_id)
            success += 1
        except Exception:
            failed += 1

    with db.connect() as conn:
        conn.execute("update import_files set success_rows = ?, failed_rows = ? where id = ?", (success, failed, file_id))

    return {"source_file_id": file_id, "success_rows": success, "failed_rows": failed}


@app.post("/api/records")
async def create_record(record: ManualRecord, authorization: str | None = Header(default=None), x_user_role: str | None = Header(default=None)) -> dict:
    user = require_login(authorization, x_user_role)
    requester = str(user.get("name") or user.get("email") or user.get("id"))
    normalized = {
        "material_name": record.material_name,
        "price": str(record.price),
        "currency": record.currency,
        "quantity": record.quantity,
        "unit": record.unit,
        "record_date": date.today().isoformat(),
        "requester": requester,
        "supplier": record.supplier,
        "remark": record.remark,
        "cas_number": record.cas_number,
    }
    record_id = await create_record_from_normalized(normalized, record.record_type, str(user.get("id")), None)
    return {"id": record_id}


async def create_record_from_normalized(normalized: dict[str, Any], record_type: str, created_by: str | None, source_file_id: int | None) -> int:
    raw_name = str(normalized.get("material_name") or normalized.get("raw_name") or "").strip()
    if not raw_name:
        raise ValueError("Missing material name.")

    candidate_obj = await resolve_chemical(str(normalized.get("cas_number") or raw_name))
    candidate = candidate_obj.__dict__ if candidate_obj else {"standard_name": raw_name, "synonyms": [raw_name], "match_status": "unresolved"}
    material_id = db.find_or_create_material(candidate, raw_name)
    quantity = parse_quantity(normalized.get("quantity") or normalized.get("quantity_value"), normalized.get("unit") or normalized.get("quantity_unit"))
    price = quantize_money(parse_decimal(normalized.get("price")))
    unit_price = quantize_money(parse_decimal(normalized.get("unit_price"))) or calculate_unit_price(price, quantity)

    with db.connect() as conn:
        cur = conn.execute(
            """
            insert into procurement_records (
                material_id, raw_name, price, currency, quantity_value, quantity_unit, normalized_quantity,
                normalized_unit, unit_price, record_type, remark, record_date, requester, supplier, source_file_id, created_by
            )
            values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                material_id,
                raw_name,
                decimal_to_str(price),
                normalized.get("currency") or "CNY",
                decimal_to_str(quantity.quantity_value),
                quantity.quantity_unit,
                decimal_to_str(quantity.normalized_quantity),
                quantity.normalized_unit,
                decimal_to_str(unit_price),
                record_type,
                normalized.get("remark"),
                normalized.get("record_date"),
                normalized.get("requester"),
                normalized.get("supplier"),
                source_file_id,
                created_by,
            ),
        )
        return int(cur.lastrowid)


@app.get("/api/materials/search")
def search_materials(q: str = "", record_type: str | None = None) -> dict:
    db.init_db()
    like = f"%{q.lower()}%"
    sql = """
        select distinct m.*
        from materials m
        left join procurement_records r on r.material_id = m.id
        where lower(m.standard_name) like ?
           or lower(coalesce(m.chinese_name, '')) like ?
           or lower(coalesce(m.english_name, '')) like ?
           or lower(coalesce(m.cas_number, '')) like ?
           or lower(m.synonyms) like ?
    """
    params: list[Any] = [like, like, like, like, like]
    if record_type:
        sql += " and r.record_type = ?"
        params.append(record_type)
    sql += " order by m.updated_at desc limit 50"
    with db.connect() as conn:
        materials = db.rows_to_dicts(conn.execute(sql, params).fetchall())
    return {"materials": materials}


@app.get("/api/materials/{material_id}/records")
def material_records(material_id: int) -> dict:
    with db.connect() as conn:
        material = db.row_to_dict(conn.execute("select * from materials where id = ?", (material_id,)).fetchone())
        rows = db.rows_to_dicts(
            conn.execute("select * from procurement_records where material_id = ? and archived_at is null order by coalesce(record_date, cast(created_at as text)) desc", (material_id,)).fetchall()
        )
        structures = db.rows_to_dicts(
            conn.execute("select * from molecule_structures where material_id = ? and archived_at is null order by updated_at desc, id desc", (material_id,)).fetchall()
        )
    if not material:
        raise HTTPException(status_code=404, detail="Material not found.")
    return {"material": material, "records": rows, "structures": structures, "stats": summarize_records(rows)}


@app.get("/api/structures")
def structure_list(q: str = "", limit: int = 100) -> dict:
    return {"structures": list_structures(q=q, limit=limit)}


@app.post("/api/structures")
def structure_register(payload: StructureRegisterRequest, authorization: str | None = Header(default=None), x_user_role: str | None = Header(default=None)) -> dict:
    user = require_login(authorization, x_user_role)
    created_by = str(user.get("email") or user.get("name") or user.get("id"))
    try:
        structure = register_structure(
            name=payload.name,
            smiles=payload.smiles,
            molfile=payload.molfile,
            cas_number=payload.cas_number,
            notes=payload.notes,
            material_id=payload.material_id,
            created_by=created_by,
        )
    except StructureError as exc:
        status_code = 409 if "已注册" in str(exc) else 400
        raise HTTPException(status_code=status_code, detail=str(exc)) from exc
    return {"structure": structure}


def structure_to_molecule(structure: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": str(structure["id"]),
        "canonical_smiles": structure.get("canonical_smiles") or "",
        "molecular_weight": structure.get("molecular_weight"),
        "molecular_formula": structure.get("molecular_formula"),
        "name": structure.get("name"),
        "notes": structure.get("notes"),
        "created_at": structure.get("created_at"),
        "updated_at": structure.get("updated_at"),
    }


def find_structure_by_canonical_smiles(smiles: str) -> dict[str, Any] | None:
    props = derive_structure_properties(smiles=smiles)
    canonical = props.canonical_smiles or smiles.strip()
    with db.connect() as conn:
        row = conn.execute(
            "select * from molecule_structures where canonical_smiles = ? and archived_at is null",
            (canonical,),
        ).fetchone()
    return db.row_to_dict(row)


def save_molecule_structure(
    *,
    smiles: str | None,
    molfile: str | None = None,
    name: str | None = None,
    notes: str | None = None,
    created_by: str = "isislike-editor",
) -> dict[str, Any]:
    try:
        return register_structure(
            name=name or smiles or "Imported molecule",
            smiles=smiles,
            molfile=molfile,
            notes=notes,
            created_by=created_by,
        )
    except StructureError as exc:
        if smiles and "已注册" in str(exc):
            existing = find_structure_by_canonical_smiles(smiles)
            if existing:
                return existing
        raise


def split_sdf_records(text: str) -> list[str]:
    return [record.strip() for record in text.split("$$$$") if record.strip()]


def parse_molecule_import(filename: str, content: bytes) -> list[dict[str, str | None]]:
    lower = filename.lower()
    if lower.endswith(".mol"):
        text = content.decode("utf-8-sig", errors="replace").strip()
        return [{"name": _clean_text(text.splitlines()[0] if text else None), "smiles": None, "molfile": text, "notes": None}]
    if lower.endswith(".sdf"):
        text = content.decode("utf-8-sig", errors="replace")
        rows = []
        for record in split_sdf_records(text):
            rows.append({"name": _clean_text(record.splitlines()[0] if record else None), "smiles": None, "molfile": record, "notes": None})
        return rows
    if lower.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip().lower() for value in next(iterator, [])]
        aliases = {
            "name": {"name", "molecule", "compound", "compound name", "chemical name", "名称", "化合物", "化合物名称"},
            "smiles": {"smiles", "canonical_smiles", "canonical smiles", "结构式", "smile"},
            "molfile": {"molfile", "mol", "mol block", "molblock"},
            "notes": {"notes", "note", "remark", "remarks", "备注"},
        }
        indexes = {
            field: next((i for i, header in enumerate(headers) if header in names), None)
            for field, names in aliases.items()
        }
        if indexes["smiles"] is None and indexes["molfile"] is None:
            indexes["smiles"] = 0
        rows = []
        for values in iterator:
            smiles = _clean_text(values[indexes["smiles"]]) if indexes["smiles"] is not None and indexes["smiles"] < len(values) else None
            molfile = _clean_text(values[indexes["molfile"]]) if indexes["molfile"] is not None and indexes["molfile"] < len(values) else None
            if not smiles and not molfile:
                continue
            name = _clean_text(values[indexes["name"]]) if indexes["name"] is not None and indexes["name"] < len(values) else None
            notes = _clean_text(values[indexes["notes"]]) if indexes["notes"] is not None and indexes["notes"] < len(values) else None
            rows.append({"name": name, "smiles": smiles, "molfile": molfile, "notes": notes})
        return rows
    raise StructureError("仅支持导入 .mol、.sdf、.xlsx 或 .xlsm 文件。")


@app.get("/api/export/config")
def export_config() -> dict:
    return {"enabled": False, "require_key": False}


@app.get("/api/molecules")
def molecule_list(limit: int = 500) -> list[dict[str, Any]]:
    return [structure_to_molecule(row) for row in list_structures(limit=limit)]


@app.post("/api/molecules/save")
def molecule_save(payload: MoleculeSaveRequest) -> dict:
    try:
        structure = save_molecule_structure(
            smiles=payload.smiles,
            molfile=payload.molfile,
            name=payload.name,
            notes=payload.notes,
        )
    except StructureError as exc:
        status_code = 409 if "已注册" in str(exc) else 400
        raise HTTPException(status_code=status_code, detail=str(exc)) from exc
    return structure_to_molecule(structure)


@app.post("/api/molecules/import")
async def molecule_import(file: UploadFile = File(...)) -> dict:
    content = await file.read()
    try:
        candidates = parse_molecule_import(file.filename or "molecules", content)
    except Exception as exc:
        detail = str(exc) or "文件无法解析。"
        raise HTTPException(status_code=400, detail=detail) from exc

    success_count = 0
    errors: list[dict[str, Any]] = []
    for index, candidate in enumerate(candidates, start=1):
        try:
            save_molecule_structure(
                name=candidate.get("name") or candidate.get("smiles") or f"Imported molecule {index}",
                smiles=candidate.get("smiles"),
                molfile=candidate.get("molfile"),
                notes=candidate.get("notes"),
                created_by="isislike-import",
            )
            success_count += 1
        except StructureError as exc:
            errors.append({"index": index, "reason": str(exc)})

    return {
        "success_count": success_count,
        "failed_count": len(errors),
        "errors": errors,
    }


@app.post("/api/molecules/search/exact")
def molecule_exact_search(payload: SmilesInput) -> dict | None:
    try:
        row = find_structure_by_canonical_smiles(payload.smiles)
    except StructureError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return structure_to_molecule(row) if row else None


@app.post("/api/molecules/search/substructure")
def molecule_substructure_search(payload: SmartsInput) -> list[dict[str, Any]]:
    return []


@app.post("/api/molecules/search/similarity")
def molecule_similarity_search(payload: SimilaritySearchRequest) -> list[dict[str, Any]]:
    return []


@app.get("/api/molecules/{molecule_id}")
def molecule_detail(molecule_id: int) -> dict:
    structure = get_structure(molecule_id)
    if not structure:
        raise HTTPException(status_code=404, detail="Molecule not found.")
    return {**structure_to_molecule(structure), "has_structure_svg": bool(structure.get("structure_svg"))}


@app.get("/api/molecules/{molecule_id}/structure.svg")
def molecule_structure_svg(molecule_id: int) -> Response:
    structure = get_structure(molecule_id)
    if not structure:
        raise HTTPException(status_code=404, detail="Structure image not found.")
    smiles = structure.get("canonical_smiles")
    if not smiles:
        raise HTTPException(status_code=404, detail="Structure SMILES not found.")
    try:
        svg = derive_structure_properties(smiles=smiles).structure_svg or render_smiles_svg(smiles)
    except StructureError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Response(content=svg, media_type="image/svg+xml")


@app.patch("/api/molecules/{molecule_id}")
def molecule_update(molecule_id: int, payload: MoleculeUpdateRequest) -> dict:
    try:
        structure = update_structure(molecule_id, name=payload.name, notes=payload.notes)
    except StructureError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return structure_to_molecule(structure)


@app.delete("/api/molecules/{molecule_id}")
def molecule_delete(molecule_id: int) -> Response:
    if not archive_structure(molecule_id):
        raise HTTPException(status_code=404, detail="Molecule not found.")
    return Response(status_code=204)


def summarize_records(records: list[dict]) -> dict:
    prices = [parse_decimal(row.get("unit_price")) for row in records if parse_decimal(row.get("unit_price")) is not None]
    latest_purchase = choose_price_source([row for row in records if row.get("record_type") == "purchase"])
    latest_inquiry = choose_price_source([row for row in records if row.get("record_type") == "inquiry"])
    return {
        "latest_purchase_unit_price": latest_purchase.get("unit_price") if latest_purchase else None,
        "latest_inquiry_unit_price": latest_inquiry.get("unit_price") if latest_inquiry else None,
        "min_unit_price": str(min(prices)) if prices else None,
        "max_unit_price": str(max(prices)) if prices else None,
        "avg_unit_price": str(sum(prices) / len(prices)) if prices else None,
    }


def calculate_quote_item(payload: QuoteRequest, persist: bool = True) -> dict:
    materials = search_materials(payload.material_query)["materials"]
    if not materials:
        return {"input": payload.model_dump(mode="json"), "material": None, "source_record": None, "result": {"estimated_cost": None, "warning": "未找到匹配物料。"}}
    material = materials[0]
    with db.connect() as conn:
        records = db.rows_to_dicts(conn.execute("select * from procurement_records where material_id = ?", (material["id"],)).fetchall())
    source = choose_price_source(records)
    result = estimate_cost(source, payload.required_quantity, payload.required_unit)
    if persist and source and result.get("estimated_cost") is not None:
        with db.connect() as conn:
            conn.execute(
                """
                insert into quotation_items (material_id, required_quantity, required_unit, latest_unit_price, estimated_cost, price_source_record_id)
                values (?, ?, ?, ?, ?, ?)
                """,
                (
                    material["id"],
                    str(payload.required_quantity),
                    payload.required_unit or "",
                    result.get("latest_unit_price"),
                    result.get("estimated_cost"),
                    result.get("price_source_record_id"),
                ),
            )
    return {"input": payload.model_dump(mode="json"), "material": material, "source_record": source, "result": result}


@app.post("/api/quotations/calculate")
def calculate_quote(payload: QuoteRequest) -> dict:
    item = calculate_quote_item(payload, persist=True)
    return {"material": item["material"], "source_record": item["source_record"], "result": item["result"]}


@app.post("/api/quotations/calculate-batch")
def calculate_quote_batch(payload: QuoteBatchRequest) -> dict:
    return {"items": [calculate_quote_item(item, persist=False) for item in payload.items]}


@app.delete("/api/quotations/{quotation_id}")
def quotation_archive(quotation_id: int, authorization: str | None = Header(default=None), x_user_role: str | None = Header(default=None)) -> dict:
    require_admin(authorization, x_user_role)
    if not archive_quotation(quotation_id):
        raise HTTPException(status_code=404, detail="Quotation not found.")
    return {"archived": True}


@app.get("/api/quotation-comparison")
def quotation_compare(left_id: int, right_id: int) -> dict:
    comparison = compare_quotations(left_id, right_id)
    if not comparison:
        raise HTTPException(status_code=404, detail="One or both quotations were not found.")
    return comparison


@app.get("/api/analytics/dashboard")
def analytics_dashboard() -> dict:
    return quotation_dashboard()


@app.post("/api/fx-rates")
def create_fx_rate(payload: FxRatePayload, authorization: str | None = Header(default=None), x_user_role: str | None = Header(default=None)) -> dict:
    user = require_admin(authorization, x_user_role)
    with db.connect() as conn:
        cur = conn.execute(
            """
            insert into fx_rates (base_currency, quote_currency, rate, effective_on, created_by)
            values (?, ?, ?, ?, ?)
            on conflict(base_currency, quote_currency, effective_on) do update set rate = excluded.rate
            """,
            (
                payload.base_currency.upper(),
                payload.quote_currency.upper(),
                str(payload.rate),
                payload.effective_on,
                user.get("id") if user else None,
            ),
        )
    return {"id": cur.lastrowid, "saved": True}


@app.get("/api/fx-rates")
def list_fx_rates() -> dict:
    with db.connect() as conn:
        rows = conn.execute("select * from fx_rates order by effective_on desc, base_currency, quote_currency").fetchall()
    return {"fx_rates": db.rows_to_dicts(rows)}


@app.post("/api/pdf/import")
async def import_pdf(file: UploadFile = File(...), authorization: str | None = Header(default=None), x_user_role: str | None = Header(default=None)) -> dict:
    require_admin(authorization, x_user_role)
    if not settings.enable_pdf_import:
        raise HTTPException(status_code=400, detail="PDF import is disabled.")
    filename = file.filename or "quotation.pdf"
    if not filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are supported.")
    content = await file.read()
    stored = remember_stored_file(filename, content, file.content_type)
    try:
        text = extract_text_pdf(content)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"stored_file": stored, "text_preview": text[:4000], "status": "text-extracted"}


@app.get("/api/records/export.csv")
def export_records() -> StreamingResponse:
    with db.connect() as conn:
        rows = db.rows_to_dicts(
            conn.execute(
                """
                select m.cas_number, m.standard_name, r.raw_name, r.price, r.currency, r.quantity_value, r.quantity_unit,
                       r.normalized_quantity, r.normalized_unit, r.unit_price, r.record_type, r.record_date, r.supplier,
                       r.requester, r.remark
                from procurement_records r
                left join materials m on m.id = r.material_id
                order by coalesce(r.record_date, cast(r.created_at as text)) desc
                """
            ).fetchall()
        )

    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=list(rows[0].keys()) if rows else ["standard_name", "price"])
    writer.writeheader()
    writer.writerows(rows)
    buffer.seek(0)
    return StreamingResponse(iter([buffer.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=quotation_records.csv"})


@app.get("/api/history/{catalog_no}/export.xlsx")
def export_catalog_history(catalog_no: str) -> StreamingResponse:
    history = get_catalog_history(catalog_no)
    if not history:
        raise HTTPException(status_code=404, detail="No quotation history found for catalog number.")
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "History"
    headers = ["catalog_no", "product_name", "unit_price", "quantity", "unit", "customer", "quoted_on", "currency", "filename"]
    sheet.append(headers)
    for row in history:
        sheet.append([row.get(header) for header in headers])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={catalog_no}_history.xlsx"},
    )
