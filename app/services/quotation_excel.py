from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import io
from typing import Any

from openpyxl import load_workbook


HEADER_ALIASES = {
    "catalog_no": ["catalogno", "catalog no", "catalog_no", "catalog number", "cat no", "货号", "编号"],
    "product_name": ["molname", "product name", "product", "name", "品名", "产品名称", "化合物名称"],
    "unit_price": ["unit price", "price", "quoted price", "单价", "报价"],
    "quantity": ["qty", "quantity", "数量", "规格"],
    "unit": ["unit", "单位"],
    "notes": ["notes", "note", "remark", "remarks", "备注", "说明"],
}

META_SHEET_NAMES = {"meta", "info", "summary"}
META_ALIASES = {
    "customer": ["customer", "client", "客户", "客户名称"],
    "quoted_on": ["quoted_on", "quote date", "quoted date", "date", "报价日期", "日期"],
    "currency": ["currency", "币种", "货币"],
}


@dataclass(frozen=True)
class ParsedQuotation:
    metadata: dict[str, str | None]
    line_items: list[dict[str, Any]]
    skipped_rows: int


def normalize_label(value: object) -> str:
    return str(value or "").strip().lower().replace("_", " ").replace("-", " ")


def parse_number(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(Decimal(str(value).strip().replace(",", "")))
    except (InvalidOperation, ValueError):
        return None


def as_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime | date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def map_headers(headers: list[object]) -> dict[int, str]:
    mapped: dict[int, str] = {}
    used_fields: set[str] = set()
    for index, header in enumerate(headers):
        normalized = normalize_label(header)
        for field, aliases in HEADER_ALIASES.items():
            if field in used_fields:
                continue
            if normalized in {normalize_label(alias) for alias in aliases}:
                mapped[index] = field
                used_fields.add(field)
                break
    return mapped


def find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[int, str]]:
    best_index = -1
    best_mapping: dict[int, str] = {}
    for index, row in enumerate(rows[:20]):
        mapping = map_headers(list(row))
        if len(mapping) > len(best_mapping):
            best_index = index
            best_mapping = mapping
        mapped_fields = set(mapping.values())
        if "catalog_no" in mapped_fields and "unit_price" in mapped_fields:
            return index, mapping
    if "catalog_no" not in set(best_mapping.values()):
        raise ValueError("Missing required catalog number column.")
    return best_index, best_mapping


def parse_meta_sheet(rows: list[tuple[Any, ...]]) -> dict[str, str | None]:
    metadata: dict[str, str | None] = {}
    for row in rows:
        if len(row) < 2:
            continue
        key = normalize_label(row[0])
        value = as_text(row[1])
        for field, aliases in META_ALIASES.items():
            if key in {normalize_label(alias) for alias in aliases}:
                metadata[field] = value
    return metadata


def parse_quotation_workbook(content: bytes) -> ParsedQuotation:
    workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    metadata: dict[str, str | None] = {}
    data_sheet = workbook.active

    for sheet in workbook.worksheets:
        if normalize_label(sheet.title) in META_SHEET_NAMES:
            metadata.update(parse_meta_sheet(list(sheet.iter_rows(values_only=True))))
        else:
            data_sheet = sheet
            break

    rows = list(data_sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Workbook does not contain quotation rows.")

    header_index, header_mapping = find_header_row(rows)
    line_items: list[dict[str, Any]] = []
    skipped_rows = 0

    for row in rows[header_index + 1 :]:
        if not any(cell not in (None, "") for cell in row):
            skipped_rows += 1
            continue
        item = {field: row[index] if index < len(row) else None for index, field in header_mapping.items()}
        catalog_no = as_text(item.get("catalog_no"))
        if not catalog_no:
            skipped_rows += 1
            continue
        line_items.append(
            {
                "catalog_no": catalog_no,
                "product_name": as_text(item.get("product_name")),
                "unit_price": parse_number(item.get("unit_price")),
                "quantity": parse_number(item.get("quantity")),
                "unit": as_text(item.get("unit")),
                "notes": as_text(item.get("notes")),
            }
        )

    return ParsedQuotation(metadata=metadata, line_items=line_items, skipped_rows=skipped_rows)
