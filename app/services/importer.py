from __future__ import annotations

import csv
import io
from decimal import Decimal
from typing import Any

from app.services.chemical_resolver import resolve_chemical
from app.services.field_mapping import detect_field_mapping, missing_required_fields
from app.services.units import calculate_unit_price, parse_decimal, parse_quantity


def _read_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


def _read_xlsx(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or "").strip() for cell in rows[0]]
    output = []
    for row in rows[1:]:
        output.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    return output


def read_table(filename: str, content: bytes) -> list[dict[str, Any]]:
    if filename.lower().endswith(".csv"):
        return _read_csv(content)
    if filename.lower().endswith(".xlsx"):
        return _read_xlsx(content)
    raise ValueError("Only .xlsx and .csv files are supported.")


def apply_mapping(row: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for header, field in mapping.items():
        normalized[field] = row.get(header)
    return normalized


async def build_preview(filename: str, content: bytes, explicit_mapping: dict[str, str] | None = None, limit: int = 50) -> dict:
    rows = read_table(filename, content)
    headers = list(rows[0].keys()) if rows else []
    mapping = explicit_mapping or detect_field_mapping(headers)
    missing = missing_required_fields(mapping)
    preview_rows = []

    for index, row in enumerate(rows[:limit], start=1):
        normalized = apply_mapping(row, mapping)
        raw_name = str(normalized.get("material_name") or "").strip()
        candidate = await resolve_chemical(raw_name)
        quantity = parse_quantity(normalized.get("quantity"), normalized.get("unit"))
        price = parse_decimal(normalized.get("price"))
        unit_price = calculate_unit_price(price, quantity)
        preview_rows.append(
            {
                "row_number": index,
                "raw": row,
                "normalized": {
                    **normalized,
                    "price": str(price) if isinstance(price, Decimal) else None,
                    "quantity_value": str(quantity.quantity_value) if quantity.quantity_value is not None else None,
                    "quantity_unit": quantity.quantity_unit,
                    "normalized_quantity": str(quantity.normalized_quantity) if quantity.normalized_quantity is not None else None,
                    "normalized_unit": quantity.normalized_unit,
                    "unit_price": str(unit_price) if unit_price is not None else None,
                },
                "material_candidate": candidate.__dict__ if candidate else None,
                "match_status": "suggested" if candidate else "unresolved",
            }
        )

    return {
        "filename": filename,
        "headers": headers,
        "mapping": mapping,
        "missing_required_fields": missing,
        "total_rows": len(rows),
        "preview_rows": preview_rows,
    }
